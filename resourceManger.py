import psutil
import openpyxl
import datetime
import time 
from web3 import Web3, HTTPProvider
class ResourceManager:

	def askHelp():
		# for now lets use the development server on ganache
		web = 'http://127.0.0.1:8545'
		# Get the client instance to interact with our ethereum network
		self.web3 = Web3(HTTPProvider(web))
		# set the default account for prototype
		self.web3.eth.defaultAccount = self.web3.eth.accounts[0]

		for account in self.web3.eth.accounts:
			if account == self.web3.eth.accounts[0]:
				continue
			tx = {
			'nonce':nonce,
			'to': account,
			'value': web3.toWei(1, 'ether')
			'gas' : 21000,
			'gasPrice': web3.toWei('50', 'gwei'),
			'data': 0x68656c70
			}

			# sign the transaction with our account
			signed_tx = web3.eth.account.signTransaction(tx, private_key)
			tx_hash = web3.eth.sendRawTransaction(signed_tx.rawTransaction)


	def agreeToHelp(account):
		web = 'http://127.0.0.1:8545'
		# Get the client instance to interact with our ethereum network
		self.web3 = Web3(HTTPProvider(web))
		# set the default account for prototype
		self.web3.eth.defaultAccount = self.web3.eth.accounts[0]

		tx = {
			'nonce':nonce,
			'to': account,
			'value': web3.toWei(1, 'ether')
			'gas' : 21000,
			'gasPrice': web3.toWei('50', 'gwei'),
			'data': 0x616363657074
			}

		signed_tx = web3.eth.account.signTransaction(tx, private_key)
		tx_hash = web3.eth.sendRawTransaction(signed_tx.rawTransaction)
		
	def data_logging(pid):
		currTime = datetime.datetime.now().strftime("%d%m%Y - %H:%M:%S")

		processP = psutil.Process(pid)

		memusage = processP.memory_percent()

		cpuTotal = processP.cpu_percent(interval=2)
		cpuUsage = processP.cpu_percent(interval=2) / psutil.cpu_count()


		resultsXL = openpyxl.load_workbook("./results.xlsx")
		sheet = resultsXL.active 

		sheet.cell(column = 1, row=sheet.max_row +1, value=currTime)
		sheet.cell(column = 2, row=sheet.max_row +1, value=pid)
		sheet.cell(column = 3, row=sheet.max_row +1, value=memusage)
		sheet.cell(column = 4, row=sheet.max_row +1, value=cpuTotal)
		sheet.cell(column = 5, row=sheet.max_row +1, value=cpuUsage)

		print('Entered data: ', currTime, pid, memusage, cpuTotal, cpuUsage)

	def monitor():
		pid = int(input('Process to be monitored: '))

		while True:
			cpuUsage = psutil.cpu_percent(interval=1)
			memUsage = psutil.virtual_memory().percent

			if cpuUsage > 75 or memUsage > 75:
				ResourceManager.askHelp()

			ResourceManager.data_logging(pid)
			time.sleep(5)



